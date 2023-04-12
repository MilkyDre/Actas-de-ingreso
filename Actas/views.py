import io
import json
import matplotlib
matplotlib.use('Agg')
from django.shortcuts import render
from django.shortcuts import render, redirect, get_object_or_404
from matplotlib import colors
from .models import *
from .forms import ingresoForm, departamentoForm, equipoForm,funcionarioForm, tecnicoForm
from django.core.mail import send_mail
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from multiselectfield import MultiSelectField
from django.urls import reverse
from rest_framework import filters
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from django.contrib import messages
import matplotlib.pyplot as plt
from django.db.models import Count
from io import BytesIO
import base64
from openpyxl import Workbook
from django.http.response import HttpResponse
from  PIL import Image
import qrcode
import os
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from openpyxl.styles import Font, Alignment, PatternFill
from django.views.generic import DetailView
import numpy as np
from datetime import timedelta
from django.utils import timezone
from django.contrib.auth.decorators import user_passes_test
from django.core.files import File

# Create your views here.
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            return render(request, 'login.html', {'error': 'Usuario o contraseña incorrectos'} )
    return render(request, 'login.html')

@login_required
def mantenimiento(request):
    return render (request,'mantenimiento.html')

@login_required
def tablas(request):
    return render (request,'tablas.html')

@login_required
def home(request):
    ingresos = Ingreso.objects.all()
    equipos = Equipo.objects.all()
    funcionarios = Funcionario.objects.all()
    en_espera_count = Ingreso.objects.filter(estado_equipo='En espera').count()
    en_reparacion_count = Ingreso.objects.filter(estado_equipo='En reparacion').count()
    reparados_count = Ingreso.objects.filter(estado_equipo='Reparado').count()
    ingresos_en_espera = Ingreso.objects.filter(estado_equipo='En espera')
    ingresos_en_reparacion = Ingreso.objects.filter(estado_equipo='En reparación')
    ingresos_reparados = Ingreso.objects.filter(estado_equipo='Reparado')

    context = {
        'ingresos': ingresos,
        'equipos': equipos,
        'funcionarios': funcionarios,
        'en_espera_count': en_espera_count,
        'en_reparacion_count': en_reparacion_count,
        'reparados_count': reparados_count,
        'ingresos_en_espera': ingresos_en_espera,
        'ingresos_en_reparacion' : ingresos_en_reparacion,
        'ingresos_reparados': ingresos_reparados,
    }
 

# Crear los datos del gráfico de pie
   # Crear el gráfico de pie
    labels = ['En espera', 'En reparacion', 'Reparado']
    sizes = [en_espera_count, en_reparacion_count, reparados_count]
    colors = ['tomato', 'orange', 'olivedrab']
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', shadow=True, startangle=90)
    ax1.axis('equal')  # para que el gráfico sea circular
    plt.tight_layout()

    # Convertir el gráfico en una imagen en BytesIO y codificarlo en base64
    buffer = io.BytesIO()
    fig1.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()
    graphic = base64.b64encode(image_png).decode('utf-8')

    # Agregar la variable graphic al diccionario de contexto
    context['graphic'] = graphic
    # Agregar el gráfico al diccionario de contexto
    return render(request, 'home.html', context)

def logout_view(request):
    logout(request)
    return redirect('login')

#------------DASHBOARD----------------#
@login_required
def export_to_excel(request):
    try:
        # crea un nuevo libro de excel
        ingresos = Ingreso.objects.all()
        
        # Crea un nuevo libro de Excel y selecciona la primera hoja
        libro = Workbook()
        hoja = libro.active

        # Define las propiedades de estilo
        cabecera_estilo = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cabecera_relleno = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        contenido_estilo = Font(name='Arial', size=11)
        contenido_alineacion = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Agrega las cabeceras a la hoja con estilo personalizado
        hoja.append(['ID', 'Técnico', 'Descripción del problema', 'Estado del equipo', 'Periféricos', 'Fecha de ingreso', 'Equipo', 'Funcionario', 'Fecha de salida', 'Descripción de reparación'])
        for celda in hoja[1]:
            celda.font = cabecera_estilo
            celda.fill = cabecera_relleno
            celda.alignment = contenido_alineacion

        # Agrega los datos de la tabla a la hoja con estilo personalizado
        for ingreso in ingresos:
            fecha_ingreso_formateada = ingreso.fecha_ingreso.strftime('%d/%m/%Y %H:%M:%S') if ingreso.fecha_ingreso else ''
            fecha_salida_formateada = ingreso.fecha_salida.strftime('%d/%m/%Y %H:%M:%S') if ingreso.fecha_salida else ''
            perifericos_seleccionados = ingreso.get_perifericos_display()
            hoja.append([ingreso.id_ingreso, ingreso.tecnico.nombre, ingreso.descripcion_problema, ingreso.estado_equipo, perifericos_seleccionados, fecha_ingreso_formateada, ingreso.equipo.serial, ingreso.funcionario.nombre, fecha_salida_formateada, ingreso.descripcion_reparacion])
            for celda in hoja[1]:
                celda.font = contenido_estilo
                celda.alignment = contenido_alineacion

                # Define el color de fondo del campo "Estado del equipo"
                if celda.column == 4:  # "Estado del equipo" es la quinta columna
                    if celda.value == 'Reparado':
                        celda.fill = PatternFill(start_color='009900', end_color='009900', fill_type='solid')  # verde
                    elif celda.value == 'En reparacion':
                        celda.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # amarillo
                    elif celda.value == 'No reparado':
                        celda.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # rojo
                    else:
                        celda.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # gris

        # Ajusta el ancho de las columnas de manera automática
        for columna in hoja.columns:
            max_length = 0
            columna = columna[0].column_letter
            for celda in hoja[columna]:
                try:
                    if len(str(celda.value)) > max_length:
                        max_length = len(celda.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            print(columna)
            hoja.column_dimensions[columna].width = adjusted_width

        # Configura la respuesta del servidor HTTP
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=Ingresos.xlsx'
        libro.save(response)
        return response

    except Exception as e:
        print(e)
        return HttpResponse("Error al exportar a Excel." + str(e))

#--------------/Equipos/----------------#
@login_required
def equipos_list(request):
    equipos = Equipo.objects.all()
    return render(request, 'equipos_list.html', {'equipos': equipos})

@login_required
def equipo_detail(request, pk):
    equipo = Equipo.objects.get(serial=pk)

    # Generar el código QR para el objeto Equipo
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(str(equipo))
    qr.make(fit=True)

    # Convertir la imagen del código QR en formato base64
    qr_img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    qr_img.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

    return render(request, 'equipo_detail.html', {'equipo': equipo, 'qr_base64': qr_base64})

def equipos_create(request):
    if request.method == 'POST':
        form = equipoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('equipos_list')
    else:
        form = equipoForm()
    return render(request, 'equipos_create.html', {'form': form})

@login_required
def equipos_edit(request, id):
    equipo = get_object_or_404(Equipo, serial=id)
    if request.method == 'POST':
        form = equipoForm(request.POST, instance=equipo)
        if form.is_valid():
            form.save()
            return redirect('equipos_list')
    else:
        form = equipoForm(instance=equipo)
    return render(request, 'equipos_edit.html', {'form': form, 'equipo': equipo})

@login_required
def equipos_delete(request, id):
    equipo = get_object_or_404(Equipo, serial=id)
    equipo.delete()
    return redirect('equipos_list')

#--------------/funcionarios/----------------#
@login_required
def funcionarios_list(request):
    funcionarios = Funcionario.objects.all()
    search_query = request.GET.get('search', '')
    funcionarios = Funcionario.objects.filter(nombre__icontains=search_query)
    context = {'funcionarios': funcionarios}
    return render(request, 'funcionarios_list.html', context)

@login_required
def funcionarios_create(request):
    if request.method == 'POST':
        form = funcionarioForm(request.POST)
        if form.is_valid():
                form.save()
                return redirect('funcionarios_list')
    else:
        form = funcionarioForm()
    return render(request, 'funcionarios_create.html', {'form': form})

@login_required
def funcionarios_edit(request, id):
    funcionario = get_object_or_404(Funcionario, rut_funcionario=id)
    if request.method == 'POST':
        form = funcionarioForm(request.POST, instance=funcionario)
        if form.is_valid():
            form.save()
            return redirect('funcionarios_list')
    else:
        form = funcionarioForm(instance=funcionario)
    return render(request, 'funcionarios_edit.html', {'form': form, 'funcionario': funcionario})

@login_required
def funcionarios_delete(request, id):
    funcionario = get_object_or_404(Funcionario, rut_funcionario=id)
    if request.method == 'POST':
        funcionario.delete()
        return redirect('funcionarios_list')
    return render(request, 'funcionarios_list.html', {'funcionario': funcionario})

#--------------/Tecnicos/----------------#
@login_required
def tecnicos_list(request):
    tecnicos = Tecnico.objects.all()
    search_query = request.GET.get('search', '')
    tecnicos = Tecnico.objects.filter(nombre__icontains=search_query)
    context = {'tecnicos': tecnicos}
    return render(request, 'tecnicos_list.html', context)

@login_required
def tecnicos_create(request):
    if request.method == 'POST':
        form = tecnicoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tecnicos_list')
    else:
        form = tecnicoForm()
    return render(request, 'tecnicos_create.html', {'form': form})

@login_required
def tecnicos_edit(request, id):
    tecnico = get_object_or_404(Tecnico, rut_tecnico=id)
    if request.method == 'POST':
        form = tecnicoForm(request.POST, instance=tecnico)
        if form.is_valid():
            form.save()
            return redirect('tecnicos_list')
    else:
        form = tecnicoForm(instance=tecnico)
    return render(request, 'tecnicos_edit.html', {'form': form, 'tecnico': tecnico})

@login_required
def tecnicos_delete(request, id):
    tecnico = get_object_or_404(Tecnico, rut_tecnico=id)
    if request.method == 'POST':
        tecnico.delete()
        return redirect('tecnicos_list')
    return render(request, 'tecnicos_list.html', {'tecnico': tecnico})

#--------------/Departamentos/----------------#
@login_required
def departamentos_list(request):
    departamentos = Departamento.objects.all()
    context = {'departamentos': departamentos}
    return render(request, 'departamentos_list.html', context)

def departamentos_search(request):
    query = request.GET.get('q')
    if query:
        departamento = Departamento.objects.filter(
            Q(id_departamento__icontains=query) |
            Q(nombre_departamento__icontains=query) 

        )
    else:
        departamento = Departamento.objects.all()
    return render(request, 'departamentos_list.html', {'departamento': departamento})

@login_required
def departamentos_create(request):
    form = departamentoForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('departamentos_list')
    return render(request, 'departamentos_create_modal.html', {'form': form})

@login_required
def departamentos_edit(request, id):
    departamento = get_object_or_404(Departamento, id_departamento=id)
    form = departamentoForm(request.POST or None, instance=departamento)
    if form.is_valid():
        form.save()
        return redirect('departamentos_list')
    return render(request, 'departamentos_edit.html', {'form': form})

@login_required
def departamentos_delete(request, id):
    departamento = get_object_or_404(Departamento, id_departamento=id)
    departamento.delete()
    return redirect('departamentos_list')

#--------------/Ingresos/----------------
@login_required
def ingresos_list(request):
    ingresos = Ingreso.objects.all()
    return render(request, 'ingresos_list.html', {'ingresos': ingresos})

@login_required
def ingresos_search(request):
    query = request.GET.get('q')
    if query:
        ingresos = Ingreso.objects.filter(
            Q(id_ingreso__icontains=query) |
            Q(tecnico__icontains=query) |
            Q(descripcion_problema__icontains=query) |
            Q(estado_equipo__icontains=query) |
            Q(perifericos__icontains=query) |
            Q(fecha_ingreso__icontains=query) |
            Q(equipo__serial__icontains=query) |
            Q(funcionario__nombre__icontains=query) 
        )
    else:
        ingresos = Ingreso.objects.all()
    return render(request, 'ingresos_list.html', {'ingresos': ingresos})

@login_required
def ingresos_create(request):
    if request.method == 'POST':
        form = ingresoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ingresos_list')
    else:
        form = ingresoForm()
    return render(request, 'ingresos_create.html', {'form': form})

@login_required    
def ingresos_update(request, id):
    ingreso = Ingreso.objects.get(id_ingreso=id)
    if request.method == 'POST':
        form = ingresoForm(request.POST, instance=ingreso)
        if form.is_valid():
            form.save()
            return redirect('ingresos_list')
    else:
        form = ingresoForm(instance=ingreso)
    return render(request, 'ingresos_edit.html', {'form': form})

@login_required
def ingresos_delete(request, id):
    ingreso = get_object_or_404(Ingreso, id_ingreso=id)
    if request.method == 'GET':
        ingreso.delete()
        return redirect('ingresos_list')
    return render(request, 'ingresos_list.html', {'ingreso': ingreso})

#--------------/CAMBIAR ESTADO/----------------#
@login_required
class IngresoDetailView(DetailView):
    model = Ingreso
    template_name = "ingreso_detail.html"
    
    def post(self, request, *args, **kwargs):
        # Obtener el objeto de ingreso a partir del id en la URL
        ingreso = get_object_or_404(Ingreso, pk=kwargs['pk'])
        
        # Actualizar el estado del equipo
        old_estado = ingreso.estado_equipo
        new_estado = request.POST.get('estado_equipo')
        ingreso.estado_equipo = new_estado
        ingreso.save()
        
        # Si el estado del equipo cambió, actualizar el QR
        if old_estado != new_estado:
            qr_image = ingreso.generate_qr_code()
            ingreso.codigo_qr.save(f"ingreso_{ingreso.id_ingreso}.png", qr_image, save=False)
            ingreso.save()
        
        return redirect('ingreso_detail', pk=kwargs['pk'])
#--------------/VISTA QR/----------------#

def scan_qr(request):
    ingresos = Ingreso.objects.all()
    for ingreso in ingresos:
        if not ingreso.codigo_qr:
            ingreso.generate_qr_code()
            ingreso.save()
    context = {
        'ingresos': ingresos
    }
    return render(request, 'qr_code.html', context)


#------------/GRAFICO DE INGRESOS (REVISAR)/----------
def grafico_ingresos(request):
    # Obtener todos los ingresos
    ingresos = Ingreso.objects.all()

    # Crear arreglos de fechas y conteos de ingresos
    dates = []
    counts = []
    for ingreso in ingresos:
        fecha = ingreso.fecha_ingreso.date()
        if fecha not in dates:
            dates.append(fecha)
            counts.append(1)
        else:
            idx = dates.index(fecha)
            counts[idx] += 1

    # Crear un arreglo de fechas para el eje x del gráfico
    start_date = min(dates)
    end_date = max(dates)
    num_days = (end_date - start_date).days + 1
    date_range = [start_date + timedelta(days=x) for x in range(num_days)]

    # Convertir los arreglos a numpy arrays
    dates_arr = np.array(dates)
    counts_arr = np.array(counts)

    # Crear el gráfico
    fig, ax = plt.subplots()
    ax.plot(date_range, np.zeros(num_days), linewidth=0, alpha=0)
    ax.bar(dates_arr, counts_arr, align='center', width=0.8)

    # Configurar el eje x del gráfico
    ax.set_xlim([start_date - timedelta(days=1), end_date + timedelta(days=1)])
    ax.xaxis_date()
    fig.autofmt_xdate()

    # Guardar el gráfico en un archivo temporal
    plt.tight_layout()
    tmpfile = BytesIO()
    fig.savefig(tmpfile, format='png')
    plt.close(fig)

    # Leer el archivo temporal y enviarlo como respuesta
    tmpfile.seek(0)
    response = HttpResponse(content_type='image/png')
    response.write(tmpfile.read())

    return response 


#--------------/Graficos tecnicos/----------------#
def grafico_tecnicos(request):
    tecnicos = Ingreso.objects.values('tecnico').annotate(ingresos=Count('id_ingreso')).order_by('tecnico')
    tecnicos_dict = {t['tecnico']: t['ingresos'] for t in tecnicos}
    tecnicos_json = json.dumps(tecnicos_dict)
    return render(request, 'grafico_tecnicos.html', {'tecnicos': tecnicos_json})